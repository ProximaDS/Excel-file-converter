import os
import xlrd
import logging
import datetime
import time

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException



# set input and output directory paths
input_dir_path = input("Enter read path:")
output_dir_path = input("Enter write path: ")


# create log file with current date and time in filename
now = datetime.datetime.now()
log_file_name = os.path.join(output_dir_path, "Process.log")
logging.basicConfig(filename=log_file_name, level=logging.DEBUG)

with open(log_file_name, "w") as f:
# loop through files in input directory
    for filename in os.listdir(input_dir_path):

        # check if file is an Excel file
        if filename.endswith('.xls'):
            # load Excel file with xlrd or openpyxl, depending on the file extension
            try:
                wb = xlrd.open_workbook(os.path.join(input_dir_path, filename))
                sheet = wb.sheet_by_index(0)
                
            except InvalidFileException:
                print(f"Error: {filename} is not a valid Excel file")
                logging.info(f"Error: - {filename} - {link['href']} - {time.strftime('%Y-%m-%d %H:%M:%S')}")
                f.write(f"{'Error'} - {filename} - {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                continue
            
            # create new Excel file with .xlsx extension
            new_filename = filename.split('.')[0] + '.xlsx'
            new_wb = Workbook()
            # copy first sheet from original file to new file
            new_sheet = new_wb.active
            new_sheet.name = sheet.name
            for row_idx in range(1, sheet.nrows):
                row = []
                for col_idx in range(sheet.ncols):
                    row.append(sheet.cell_value(row_idx, col_idx))
                new_sheet.append(row)
            # save new file to output directory
            new_wb.save(os.path.join(output_dir_path, new_filename))\
            
            logging.info(f"Completed: - {new_filename} - {time.strftime('%Y-%m-%d %H:%M:%S')}")
            f.write(f"{'Completed'} - {new_filename} - {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        elif filename.endswith('.xlsx'):
            try:
                wb = load_workbook(os.path.join(input_dir_path, filename))
                sheet = wb.active
            except InvalidFileException:
                print(f"Error: {filename} is not a valid Excel file")
                logging.info(f"Error:- {filename} - {time.strftime('%Y-%m-%d %H:%M:%S')}")
                f.write(f"{'Error'} - {filename} - {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
                continue
            
            # create new Excel file with .xlsx extension
            new_filename = filename.split('.')[0] + '.xlsx'
            new_wb = Workbook()
            # copy first sheet from original file to new file
            new_sheet = new_wb.active
            new_sheet.title = sheet.title
            for row in sheet.iter_rows(values_only=True):
                new_sheet.append(row)
            # save new file to output directory
            new_wb.save(os.path.join(output_dir_path, new_filename))
            logging.info(f"Completed:- {new_filename} - {time.strftime('%Y-%m-%d %H:%M:%S')}")
            f.write(f"{'Completed'} - {new_filename} - {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        else:
            print("/n")

    time.sleep(1)
        
print("Conversion complete!")